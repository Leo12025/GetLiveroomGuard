# 获取B站直播间舰长列表工具

import requests
import argparse
import json
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, Alignment

# 设置请求头，模拟浏览器访问
HEADERS = {
    'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/110.0.0.0 Safari/537.36',
    'Referer': 'https://live.bilibili.com/',
    'Cookie': 'SESSDATA=your_sessdata_here; bili_jct=your_bili_jct_here'  # 需要替换为实际的Cookie
}

# B站API基础URL
BASE_URL = 'https://api.live.bilibili.com'

class BiliBiliLiveGuard:
    def __init__(self, room_id, ruid=None):
        self.room_id = room_id
        self.ruid = ruid
        self.guard_list = []

    def get_guard_info(self):
        """获取直播间舰长信息"""
        url = f'{BASE_URL}/xlive/app-room/v2/guardTab/topListNew'
        params = {
            'roomid': self.room_id,
            'page': 1,
            'page_size': 20,  # 每页获取20条数据
            'typ': 0
        }

        # 如果提供了ruid，则添加到参数中
        if self.ruid:
            params['ruid'] = self.ruid

        try:
            response = requests.get(url, headers=HEADERS, params=params)
            if response.status_code == 200:
                data = response.json()
                if data['code'] == 0:
                    # 处理分页
                    # 优先从data.info.page获取总页数，以适应API数据结构
                    total_page = data['data'].get('info', {}).get('page', data['data'].get('page', 1))
                    self._process_guard_data(data['data'].get('list', []))

                    # 获取后续页面数据
                    for page in range(2, total_page + 1):
                        params['page'] = page
                        page_response = requests.get(url, headers=HEADERS, params=params)
                        if page_response.status_code == 200:
                            page_data = page_response.json()
                            if page_data['code'] == 0:
                                self._process_guard_data(page_data['data'].get('list', []))
                            else:
                                print(f'获取第{page}页数据失败: {page_data.get("message", "未知错误")}')
                        else:
                            print(f'请求第{page}页失败，状态码: {page_response.status_code}')

                    return True
                else:
                    print(f'API返回错误: {data.get("message", "未知错误")}')
                    return False
            else:
                print(f'请求失败，状态码: {response.status_code}')
                return False
        except Exception as e:
            print(f'发生异常: {str(e)}')
            return False

    def _process_guard_data(self, guard_data):
        """处理舰长数据"""
        for guard in guard_data:
            # 提取所需信息
            user_info = {
                '用户名': guard.get('uinfo', {}).get('base', {}).get('name', '未知'),
                'uid': guard.get('uinfo', {}).get('uid', 0),
                '舰长等级': self._get_guard_level(guard.get('uinfo', {}).get('guard', {}).get('level', 0)),
                '勋章等级': guard.get('uinfo', {}).get('medal', {}).get('level', 0),
                '30天消费': guard.get('accompany', 0),  # 30天消费
                '排名': guard.get('rank', 0),
                '头像URL': guard.get('uinfo', {}).get('base', {}).get('face', '').strip()  # 移除多余的空格
            }
            self.guard_list.append(user_info)

    def _get_guard_level(self, level):
        """将数字等级转换为文字"""
        levels = {1: '总督', 2: '提督', 3: '舰长'}
        return levels.get(level, f'未知等级({level})')

    def print_guard_list(self):
        """打印舰长列表"""
        if not self.guard_list:
            print('未获取到舰长信息')
            return

        print(f'直播间 {self.room_id} 的舰长列表 ({len(self.guard_list)}人):')
        print('-' * 120)
        # 格式化输出表头
        print(f'{"用户名":<20}{"UID":<15}{"舰长等级":<10}{"勋章等级":<10}{"排名":<8}{"30天消费":<10}{"头像URL"}')
        print('-' * 120)

        # 打印每个舰长的信息
        for guard in self.guard_list:
            print(f'{guard["用户名"]:<20}{guard["uid"]:<15}{guard["舰长等级"]:<10}{guard["勋章等级"]:<10}{guard["排名"]:<8}{guard["30天消费"]:<10}{guard["头像URL"]}')

    def save_to_file(self, filename=None):
        """保存舰长列表到JSON文件"""
        if not self.guard_list:
            print('没有舰长信息可保存')
            return False

        if not filename:
            filename = f'guard_list_{self.room_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.json'

        try:
            with open(filename, 'w', encoding='utf-8') as f:
                json.dump(self.guard_list, f, ensure_ascii=False, indent=2)
            print(f'舰长信息已保存到 {filename}')
            return True
        except Exception as e:
            print(f'保存文件失败: {str(e)}')
            return False

    def save_to_excel(self, filename=None):
        """保存舰长列表到Excel文件"""
        if not self.guard_list:
            print('没有舰长信息可保存')
            return False

        if not filename:
            filename = f'guard_list_{self.room_id}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'

        try:
            # 创建工作簿和工作表
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = '舰长列表'

            # 设置表头
            headers = ['用户名', 'UID', '舰长等级', '勋章等级', '排名', '30天消费', '头像URL']
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_idx)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # 填充数据
            for row_idx, guard in enumerate(self.guard_list, 2):
                ws.cell(row=row_idx, column=1).value = guard['用户名']
                ws.cell(row=row_idx, column=2).value = guard['uid']
                ws.cell(row=row_idx, column=3).value = guard['舰长等级']
                ws.cell(row=row_idx, column=4).value = guard['勋章等级']
                ws.cell(row=row_idx, column=5).value = guard['排名']
                ws.cell(row=row_idx, column=6).value = guard['30天消费']
                ws.cell(row=row_idx, column=7).value = guard['头像URL']

            # 调整列宽
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter  # 获取列字母
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2) * 1.2
                ws.column_dimensions[column].width = adjusted_width

            # 保存文件
            wb.save(filename)
            print(f'舰长信息已保存到Excel文件: {filename}')
            return True
        except Exception as e:
            print(f'保存Excel文件失败: {str(e)}')
            return False

def test_with_sample_data():
    """使用示例数据测试程序"""
    import os
    # 创建实例
    bili_guard = BiliBiliLiveGuard(12345)  # 使用任意房间ID

    # 读取示例数据
    sample_file = 'exp.json'
    if os.path.exists(sample_file):
        try:
            with open(sample_file, 'r', encoding='utf-8') as f:
                data = json.load(f)
                if data['code'] == 0:
                    # 处理示例数据
                    bili_guard._process_guard_data(data['data']['list'])
                    print(f'成功处理示例数据，共 {len(bili_guard.guard_list)} 条记录')
                    
                    # 打印舰长列表
                    bili_guard.print_guard_list()
                    
                    # 保存到Excel
                    bili_guard.save_to_excel('sample_guard_list.xlsx')
                    return True
                else:
                    print(f'示例数据格式错误: {data.get("message", "未知错误")}')
                    return False
        except Exception as e:
            print(f'读取示例数据失败: {str(e)}')
            return False
    else:
        print(f'未找到示例数据文件: {sample_file}')
        return False

def main():
    # 解析命令行参数
    parser = argparse.ArgumentParser(description='获取B站直播间舰长列表')
    parser.add_argument('room_id', type=int, help='直播间ID')
    parser.add_argument('ruid', type=int, help='主播UID')
    parser.add_argument('-o', '--output', help='输出文件名')
    parser.add_argument('-e', '--excel', action='store_true', help='导出为Excel格式')
    parser.add_argument('-t', '--test', action='store_true', help='使用示例数据测试')
    args = parser.parse_args()

    if args.test:
        test_with_sample_data()
    elif args.room_id:
        # 创建实例并获取舰长信息
        bili_guard = BiliBiliLiveGuard(args.room_id, args.ruid)
        if bili_guard.get_guard_info():
            # 打印舰长列表
            bili_guard.print_guard_list()

            # 保存到文件
            if args.output:
                if args.excel:
                    bili_guard.save_to_excel(args.output)
                else:
                    bili_guard.save_to_file(args.output)
            elif args.excel:
                bili_guard.save_to_excel()
    else:
        parser.error('必须提供直播间ID和主播UID，或使用-t选项进行测试')

if __name__ == '__main__':
    main()

# 使用说明:
# 1. 替换代码中的Cookie为你自己的B站Cookie
# 2. 运行命令: python3 get_liveroom_guard.py 直播间ID 主播UID
# 3. 可选参数:
#    -o 文件名 保存舰长信息到指定文件
#    -e 导出为Excel格式 (与-o参数一起使用时将保存为Excel文件)
#    -t 使用示例数据测试，无需提供直播间ID和主播UID